import openpyxl
import json
import datetime
import os

excel_path = r"C:\Users\DELL\Downloads\KINDNET_Sample_Data.xlsx"
json_path = r"c:\laragon\www\kindnetV2\database\seeders\kindnet_sample_data.json"

def parse_date(val, name=""):
    if not val:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime('%Y-%m-%d')
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ('none', 'null', ''):
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.datetime.strptime(val_str, fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    print(f"Warning: Could not parse date value '{val_str}' for column '{name}'")
    return val_str

def parse_numeric(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    val_str = str(val).strip().replace(',', '')
    if not val_str or val_str.lower() in ('none', 'null', ''):
        return None
    try:
        if '.' in val_str:
            return float(val_str)
        return int(val_str)
    except ValueError:
        return val_str

def clean_value(val, col_name):
    if val is None:
        return None
    val_str = str(val).strip()
    if val_str.lower() in ('none', 'null', ''):
        return None
    
    # Check if this column is a date
    if 'date' in col_name or 'birth' in col_name or 'dob' in col_name:
        return parse_date(val, col_name)
    
    # Check if this column is numeric
    if any(suffix in col_name for suffix in ['_rm', 'income', 'dependents', 'year', 'number']):
        # Except bank account number, which should be kept as string
        if col_name == 'account_number':
            return val_str
        return parse_numeric(val)
    
    return val_str

def read_sheet(sheet):
    rows = list(sheet.rows)
    if not rows:
        return []
    
    headers = [c.value for c in rows[0]]
    data = []
    
    for row in rows[1:]:
        row_data = {}
        for h, cell in zip(headers, row):
            if h is None:
                continue
            row_data[h] = clean_value(cell.value, h)
        # Check if row is empty
        if any(v is not None for v in row_data.values()):
            data.append(row_data)
            
    return data

def main():
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        return
        
    print(f"Loading workbook from {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    result = {}
    
    if 'Applicants' in wb.sheetnames:
        print("Reading Applicants sheet...")
        result['applicants'] = read_sheet(wb['Applicants'])
    else:
        print("Warning: Applicants sheet not found")
        
    if 'Applications' in wb.sheetnames:
        print("Reading Applications sheet...")
        result['applications'] = read_sheet(wb['Applications'])
    else:
        print("Warning: Applications sheet not found")
        
    print(f"Writing parsed data to {json_path}...")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, indent=4, ensure_ascii=False)
        
    print(f"Done! Successfully wrote {len(result.get('applicants', []))} applicants and {len(result.get('applications', []))} applications.")

if __name__ == '__main__':
    main()
